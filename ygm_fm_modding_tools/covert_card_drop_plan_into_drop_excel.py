import openpyxl.workbook
import polars as pl
from polars import selectors as cs
import openpyxl
from pathlib import Path
import logging


def card_drop_plan_to_card_drop_proba(card_drop_plan: pl.DataFrame) -> pl.DataFrame:
    """
    Convert a card drop plan specifying only the droppable cards of each oppenents for each rank to dataframe with probabilities of drops of each card (each card have almost same probabilities to drop for a same rank of a same character)

    Parameters
    ----------
    card_drop_plan : pl.DataFrame
        card drop plan specifying only the droppable cards of each oppenents for each rank

    Returns
    -------
    pl.DataFrame
        dataframe with probabilities of drops of each card (each card have almost same probabilities to drop for a same rank of a same character)
    """
    # Extract the first row
    first_row = card_drop_plan[0].to_dict(as_series=False)
    # Rename columns by merging original column names with first row values
    new_columns = [f"{col}_{first_row[col][0]}" for col in card_drop_plan.columns]
    # new_columns
    # # Apply the new column names to the DataFrame
    card_drop_plan = card_drop_plan.rename(
        dict(zip(card_drop_plan.columns, new_columns))
    )

    # # Drop the first row
    card_drop_plan = card_drop_plan.slice(1)

    # df
    card_drop_plan = (
        card_drop_plan.rename(
            {"__UNNAMED__0_None": "card_id", "24.06666667_146.2857143": "card_name"}
        )
        .filter(~pl.col("card_id").is_null())
        .drop("409.6_Total occurence")
        .drop(cs.contains("_None"))
        .with_columns((~cs.contains("card_")).cast(int))
        .fill_null(0)
        .with_columns((~cs.contains("card_")) * (2048 // (~cs.contains("card_")).sum()))
    )
    col_selector = ~cs.contains("card_")
    columns_to_modify = card_drop_plan.select(col_selector).columns
    for col in columns_to_modify:
        card_drop_plan = card_drop_plan.with_columns(
            pl.when(
                (pl.col(col) != 0) & (pl.col(col).replace(0, None).cum_count() == 1)
            )
            .then(pl.col(col) + ((2048 - pl.col(col).sum())))
            .otherwise(
                pl.when((pl.col(col) != 0)).then(pl.col(col)).otherwise(pl.lit(0))
            )
            .alias(col)
        )
    return card_drop_plan


def card_drop_proba_to_drop_template_excel(
    card_drop_proba: pl.DataFrame, card_drop_excel_template: openpyxl.workbook
) -> None:

    sheet = card_drop_excel_template.active
    column_names = [
        cell.value for cell in sheet[4]
    ]  # Assuming the first row contains headers

    opponents = [name for name in column_names if name is not None]

    opponents.remove("Oponente*")
    # Define the starting row
    start_row = 6
    print(card_drop_proba.columns)

    for opponent_index in range(len(opponents)):
        print(opponents[opponent_index])
        for rank in range(3):
            column_index = (4 * (opponent_index + 1)) + rank
            print(sheet.cell(row=5, column=column_index).value)
            proba = []
            for i, value in enumerate(
                card_drop_proba[:, (((3 * (opponent_index)) + rank) + 2)].to_list(),
                start=start_row,
            ):

                cell = sheet.cell(row=i, column=column_index, value=value)
                if value != 0:
                    proba.append(cell.value)
            print(proba)


if __name__ == "__main__":
    data_folder = Path(__file__).parent.parent / "data"
    card_drop_plan_path = data_folder / "Card_Drop_Plan.xlsx"
    df = pl.read_excel(card_drop_plan_path)
    logging.info("Convert card drop plan to card drop proba")
    card_drop_proba = card_drop_plan_to_card_drop_proba(df)
    card_drop_excel_template = openpyxl.load_workbook(
        data_folder / "Drop_Rebalanced.xlsx"
    )
    logging.info("Convert card drop proba to template")

    card_drop_proba_to_drop_template_excel(
        card_drop_proba=card_drop_proba,
        card_drop_excel_template=card_drop_excel_template,
    )
    card_drop_excel_template.save(data_folder / "update_drop_template.xlsx")
